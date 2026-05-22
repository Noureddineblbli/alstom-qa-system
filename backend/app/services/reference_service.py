import json
import os
import shutil
from sqlalchemy.orm import Session
from app.models.reference import Reference
from app.models.reference_data import ReferenceData
from app.utils.excel_parser import parse_reference_file
from app.core.config import settings
from app.models.user import User

def get_all_references(db: Session):
    results = (
        db.query(Reference, User.nom, User.email)
        .join(User, Reference.created_by == User.user_id)
        .all()
    )

    return [
        {
            "ref_id": ref.ref_id,
            "project_id": ref.project_id,
            "created_at": ref.created_at,
            "created_by": username,   # 👈 name instead of ID
            "email": email
        }
        for ref, username, email in results
    ]

def get_reference_by_id(ref_id: str, db: Session):
    return db.query(Reference).filter(Reference.ref_id == ref_id).first()

def get_reference_slots(ref_id: str, db: Session):
    return db.query(ReferenceData).filter(ReferenceData.ref_id == ref_id).all()

def create_reference(ref_id: str, project_id: int, created_by: int, row_results: str, file, db: Session):
    # Save uploaded file temporarily
    if file:
        temp_path = f"uploads/{file.filename}"
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

    try:
        # Parse the Excel/CSV file
        if file:
            slots = parse_reference_file(temp_path)
        else:
            # Parse row results from frontend
            slots_data = json.loads(row_results)
            print(f"DEBUG: raw type is {type(slots_data)}")
            print(f"DEBUG: raw value is {slots_data}")
            
            slots = []
            for row_idx, items in slots_data.items():
                for item in items:
                    slots.append({
                        "slotId": item["slot_id"],
                        "amperage": item["scanned_calibre"],
                        "identification_id": item["scanned_identification"]
                    })


        # Create the reference record
        reference = Reference(
            ref_id=ref_id,
            project_id=project_id,
            created_by=created_by
        )
        db.add(reference)
        db.commit()
        db.refresh(reference)

        # Create slot records
        for slot in slots:
            slot_record = ReferenceData(
                slotId=slot["slotId"],
                ref_id=reference.ref_id,
                identification_id=slot["identification_id"],
                amperage=slot["amperage"]
            )
            db.add(slot_record)

        db.commit()
        return reference

        #         # IMPORTANT: build frontend response
        # formatted_slots = [
        #     {
        #         "slot_id": slot["slotId"],
        #         "expected_identification": slot["identification_id"],
        #         "expected_calibre": slot["amperage"]
        #     }
        #     for slot in slots
        # ]

        # return {
        #     "id": reference.ref_id,
        #     "name": reference.ref_id,
        #     "projectId": reference.project_id,
        #     "updatedAt": reference.created_at.isoformat()
        #     if hasattr(reference, "created_at")
        #     else None,
        #     "slots": formatted_slots
        # }

    except Exception as e:
        db.rollback()
        raise ValueError(str(e))
    finally:
        # Clean up temp file
        if file:
            if os.path.exists(temp_path):
                os.remove(temp_path)

def delete_reference(ref_id: str, db: Session):
    reference = db.query(Reference).filter(Reference.ref_id == ref_id).first()
    if not reference:
        return None
    db.query(ReferenceData).filter(ReferenceData.ref_id == ref_id).delete()
    db.delete(reference)
    db.commit()
    return True

def update_reference(ref_id: str, new_project_id: int, db: Session):
    reference = db.query(Reference).filter(Reference.ref_id == ref_id).first()
    if not reference:
        return None
    reference.project_id = new_project_id
    db.commit()
    db.refresh(reference)
    return reference

def bulk_update_reference_slots(ref_id: str, slots: list, db: Session):
    updated = []
    not_found = []

    for slot_data in slots:
        slot = db.query(ReferenceData).filter(
            ReferenceData.slotId == slot_data.slotId,
            ReferenceData.ref_id == ref_id
        ).first()

        if not slot:
            not_found.append(slot_data.slotId)
            continue

        slot.identification_id = slot_data.identification_id
        slot.amperage = slot_data.amperage
        updated.append(slot)

    if not_found:
        db.rollback()
        raise ValueError(f"Slots not found: {', '.join(not_found)}")

    db.commit()
    for slot in updated:
        db.refresh(slot)

    return updated


import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def generate_reference_template() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reference Template"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    subheader_fill = PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    # Row 1: ROW label + slot group headers (S1 to S19)
    ws.cell(row=1, column=1, value="Row").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center

    col = 2
    for slot_num in range(1, 20):  # S1 to S19
        # Merge 2 columns for each slot header
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
        cell = ws.cell(row=1, column=col, value=f"S{slot_num}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        col += 2

    # Row 2: ID and AMP subheaders for each slot
    ws.cell(row=2, column=1, value="").fill = subheader_fill
    col = 2
    for slot_num in range(1, 20):
        id_cell = ws.cell(row=2, column=col, value="ID")
        id_cell.font = Font(bold=True)
        id_cell.fill = subheader_fill
        id_cell.alignment = center

        amp_cell = ws.cell(row=2, column=col+1, value="AMP")
        amp_cell.font = Font(bold=True)
        amp_cell.fill = subheader_fill
        amp_cell.alignment = center
        col += 2

    # Rows 3-7: Data rows for rows 1 to 5
    for row_num in range(1, 6):
        ws.cell(row=row_num+2, column=1, value=row_num).alignment = center

    # Column widths
    ws.column_dimensions['A'].width = 6
    for col_idx in range(2, 40):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 9

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer